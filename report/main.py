import argparse
import datetime
import getpass
import sys

import openpyxl

from report.db import Database
from report.pdf import Pdf

"""
Main module (command-line program).
"""


class ArgParser(argparse.ArgumentParser):
    """
    Argument parser that displays help on error
    """
    def error(self, message):
        self.print_help()
        sys.stderr.write("error: {}\n".format(message))
        sys.exit(2)


def _parse_arguments():
    parser = ArgParser(
        description="Demo report generator",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument("host", help="Database host IP address")
    parser.add_argument("port", help="Database host port")
    parser.add_argument("user", help="Database user")
    parser.add_argument("dbname", help="Database name")
    args = parser.parse_args()
    args.password = getpass.getpass()
    return args


def store_address(db, pdf):
    ADDRESS_SQL = """
      select s.store_id, a.address, c.city, cn.country
        from store as s
        join address as a on s.address_id = a.address_id
        join city as c on c.city_id = a.city_id
        join country as cn on cn.country_id = c.country_id
       limit 1;
    """
    address = db.execute(ADDRESS_SQL)
    pdf.set_store_address(address[0][1:])


def rentals_by_day(db, pdf, wb, start_date, end_date):
    RENTALS_BY_DAY_SQL = """
      select date_trunc('day', r.rental_date) as day, count(*) as n
        from rental as r
       where r.rental_date >= %(start)s and r.rental_date < %(end)s
    group by day
    order by day;
    """
    rentals = db.execute(RENTALS_BY_DAY_SQL,
                         dict(start=start_date, end=end_date))
    rentals_dict = {d.day: n for d, n in rentals}
    days = list(range(1, 32))
    n_rents = [rentals_dict.get(d, 0) for d in days]

    pdf.add_paragraph("Number of rentals per day in Jun 2005:")
    pdf.add_line_chart(170, 50, [str(d) for d in days], [n_rents])

    ws = wb.create_sheet("rentals")
    ws["A1"] = "Day"
    ws["B1"] = "No Rentals"
    for row in zip(days, n_rents):
        ws.append(row)


def top_ten_customers_by_count(db, pdf, wb, start_date, end_date):
    TOP_TEN_CUSTOMERS_BY_COUNT = """
      select c.first_name, c.last_name, count(c.customer_id) as n
        from customer as c
        join rental as r on r.customer_id = c.customer_id
       where r.rental_date >= %s and r.rental_date < %s
    group by c.customer_id
    order by n desc
       limit 10;
    """
    top_customers = db.execute(TOP_TEN_CUSTOMERS_BY_COUNT,
                               (start_date, end_date))
    top_customers_str = [
        (fn, ln, str(cnt)) for fn, ln, cnt in top_customers
    ]
    rows = [
        ("First Name", "Last Name", "Count"),
    ] + top_customers_str
    sizes = 60, 60, 20

    pdf.add_paragraph("Top 10 customers (by number of rentals) for Jun 2005:")
    pdf.add_table(rows, sizes)

    ws = wb.create_sheet("customers")
    ws.append(rows[0])
    for i, (fn, ln, n) in enumerate(top_customers, 2):
        ws.cell(i, 1, fn)
        ws.cell(i, 2, ln)
        ws.cell(i, 3, n)


def top_ten_actors(db, pdf, wb, start_date, end_date):
    TOP_TEN_ACTORS = """
      select a.first_name, a.last_name, count(a.actor_id) as n
        from actor as a
        join film_actor as fa on fa.actor_id = a.actor_id
        join film as f on f.film_id = fa.film_id
        join inventory as i on i.film_id = f.film_id
        join rental as r on i.inventory_id = r.inventory_id
       where r.rental_date >= %(start)s and r.rental_date < %(end)s
    group by a.actor_id
    order by n desc
       limit 10;
    """
    actors = db.execute(TOP_TEN_ACTORS,
                        {"start": start_date, "end": end_date})
    actor_labels = tuple("{} {}".format(fn, ln) for fn, ln, _ in actors)
    actor_shares = tuple(a[2] for a in actors)

    pdf.add_paragraph("Top ten actors in Jun 2005:")
    pdf.add_pie_chart(80, 80, actor_labels, actor_shares, side_labels=True)

    ws = wb.create_sheet("actors")
    ws.append(("First Name", "Last Name", "Count"))
    for a in actors:
        ws.append(a)


def main():
    args = _parse_arguments()

    db = Database(args.host, args.port, args.user, args.password, args.dbname)
    wb = openpyxl.Workbook()
    pdf = Pdf("test.pdf")

    start_date = datetime.date(2005, 6, 1)
    end_date = datetime.date(2005, 7, 1)

    store_address(db, pdf)
    rentals_by_day(db, pdf, wb, start_date, end_date)
    top_ten_customers_by_count(db, pdf, wb, start_date, end_date)
    top_ten_actors(db, pdf, wb, start_date, end_date)

    pdf.save()
    wb.save("sample.xlsx")
    return 0
